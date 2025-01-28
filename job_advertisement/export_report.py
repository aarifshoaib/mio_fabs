from django.shortcuts import render, redirect
from django.views import View
from job_advertisement import models
from datetime import datetime, timedelta
from django.contrib import messages
from client_management.models import AddCompanyModel
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font

class JobReportExportView(View):
    def get(self, request):
        response = HttpResponse(content_type='text/ms-excel')
        file_name = f"remainder60days.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        today = datetime.today().date()
        matured_jobs = models.JobAdvertisementModel.objects.filter(
                                expiry_date__gte=today).order_by('-date_of_app')
        jobs = []
        for mature in matured_jobs:
            if (mature.expiry_date-today).days<=30:
                jobs.append(mature)
        columns = ['PostDate', 'JobCode', 'Expiry', 'CompanyName']
        ws.append(columns)
        for h in range(1, 5):
            ws.cell(row=1, column=h).font = Font(bold=True)
            ws.cell(row=1, column=h).fill = openpyxl.styles.PatternFill(start_color="2a9df4",
                                                end_color="2a9df4", fill_type = 'solid')
        for job in jobs:
            v = [job.date_of_app, job.job_code, job.expiry_date, job.comp_name.company_name]
            ws.append(v)
        wb.save(response)
        return response

class CompanyReportExportView(View):
    def get(self, request):
        response = HttpResponse(content_type='text/ms-excel')
        file_name = f"companies-list.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        today = datetime.today().date()
        columns = 'ClientID', 'Client Name', 'CompanyID', 'Company Name', 'ROC', 'Phone Number'
        ws.append(columns)
        for h in range(1, 7):
            ws.cell(row=1, column=h).font = Font(bold=True)
            ws.cell(row=1, column=h).fill = openpyxl.styles.PatternFill(start_color="2a9df4",
                                                end_color="2a9df4", fill_type = 'solid')
        companies = AddCompanyModel.objects.all().values('client__client_id', 'client__client_name',
                                'company_id', 'company_name', 'roc', 'client__phone_number').order_by('client__client_id', 'company_id')
        for comp in companies:
            v = [comp['client__client_id'], comp['client__client_name'], comp['company_id'],
                            comp['company_name'], comp['roc'], comp['client__phone_number']]
            ws.append(v)
        wb.save(response)
        return response
